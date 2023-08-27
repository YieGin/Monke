import tkinter as tk
import tkinter.ttk as ttk
from genericpath import exists
import sys
import os
import warnings
from openpyxl import load_workbook
import xlsxwriter as xls
from datetime import datetime

sys.path.append("C:\\Users\\blade\\Downloads\\UPA")
sys.path.append("C:\\Users\\blade\\Downloads\\UPA\\libIBS")
import libIBS as ibs


def main(root):
    global inputFiles
    global outputFile
    global reportYear
    global popup
    global popupBar

    # Create a window
    popup = ibs.create_popup(root, "UPA rapportage", 800, 120)

    # progressbar
    popupBar = ibs.create_popup(popup, "Voortgang", 1000, 115)
    ibs.hide_window(popupBar)
    popup.focus_force()

    # input field for file name input
    tk.Label(popup, text="Invoerbestand(en):").place(x=10, y=5)
    tk.Label(
        popup, text="Bij meerdere invoerbestanden ; als delimiter gebruiken."
    ).place(x=150, y=30)
    inputFiles = tk.Entry(popup, width=100)
    inputFiles.place(x=150, y=5)
    btnSelectInputFiles = tk.Button(
        popup,
        text="...",
        command=lambda: select_files(inputFiles, "Selecteer invoerbestanden"),
    )
    btnSelectInputFiles.place(x=755, y=1)

    # input field for output file name
    tk.Label(popup, text="Uitvoer Excel-bestand:").place(x=10, y=55)
    outputFile = tk.Entry(popup, width=100)
    outputFile.place(x=150, y=55)
    btnSelectOutputFile = tk.Button(
        popup,
        text="...",
        command=lambda: saveas_filename(outputFile, "Selecteer uitvoerbestand"),
    )
    btnSelectOutputFile.place(x=755, y=55)

    # input field for report year
    tk.Label(popup, text="Rapportage jaar:").place(x=10, y=80)
    reportYear = tk.Entry(popup, width=5)
    reportYear.place(x=150, y=80)
    reportYear.insert(0, ibs.current_year())

    # Create a Button Widget in the popup Window
    btnOK = tk.Button(popup, text="Maak", width=10, command=check_input)
    btnOK.place(x=610, y=85)
    btnCancel = tk.Button(popup, text="Sluit", width=10, command=close_popup)
    btnCancel.place(x=695, y=85)


def close_popup():
    ibs.close_window(popupBar)
    ibs.close_window(popup)


def select_files(entryField, title):
    selectionList = (
        ("Excel bestanden", "*.xls;*.xlsx;*.xlsm"),
        ("Alle bestanden", "*.*"),
    )
    selectedFile = ibs.select_file(selectionList, "", title, True)
    if selectedFile is not None and selectedFile != "":
        list = []
        f = entryField.get()
        if len(f) > 0:
            list = f.split(";")
        for f in selectedFile:
            f = f.strip("{")
            f = f.strip("}")
            if f not in list:
                list.append(f)
        f = ";".join(list)
        entryField.delete(0, "end")
        entryField.insert(0, f)
    popup.focus_force()
    return


def default_output():
    return ibs.date_yyyymmdd_hhmmss() + "_UPA_foutenrapportage.xlsx"


def saveas_filename(entryField, title):
    selectionList = (("Excel-bestanden", "*.xlsx"), ("Alle bestanden", "*.*"))
    selectedFile = ibs.saveas_filename(selectionList, "", title, default_output())
    if selectedFile is not None and selectedFile != "":
        entryField.delete(0, "end")
        entryField.insert(0, selectedFile)
    popup.focus_force()
    return


def check_input():
    files = inputFiles.get()
    output = outputFile.get()
    if files.strip() == "":
        ibs.display_message("*** Fout ***", "Geen invoerbestand ingegeven!")
        return
    if output.strip() == "":
        ibs.display_message("*** Fout ***", "Geen uitvoerbestand ingegeven!")
        return

    files = files.strip(";")
    files = files.replace("\\", "/")
    inputList = files.split(";")
    maak_rapportage(inputList, output)

    return


def maak_rapportage(list, output):
    warnings.simplefilter("ignore")
    output = outputFile.get()
    basename, extension = os.path.splitext(output)

    content = []
    for f in list:
        if exists(f):
            content.extend(
                lees_bestand(f)
            )  # gecheckt en wordt goed ingelezen. Klopt volgens de onderstaande vulling

    # vulling kolommen content:
    # 0 = Maand
    # 1 = Fout
    # 2 = Omschrijving
    # 3 = SWO code
    # 4 = SWO naam
    # 5 = Aantal fouten
    # 6 = Aantal WG
    # 7 = Aantal IKV
    # 8 = Omschrijving impact
    # 9 = PUO

    process_total(content, basename, extension)
    process_swo(content, basename, extension)

    ibs.display_message("Informatie", "UPA foutenrapportage is aangemaakt.")

    return


def get_dateformat():
    return {"num_format": "mmm-yy"}


def get_headerformat():
    return {
        "bold": True,
        "text_wrap": True,
        "valign": "top",
        "font_color": "white",
        "fg_color": "black",
        "border": 0,
    }


def get_cellformat():
    return {"text_wrap": True}


def process_total(content, basename, extension):
    if len(content) > 0:
        # progressbar
        ibs.show_window(popupBar)

        bar = ttk.Progressbar(
            popupBar, orient=tk.HORIZONTAL, length=950, mode="determinate"
        )
        txt = tk.Label(popupBar, text="Maken totaaloverzicht... ")
        bar.pack(expand=True)
        bar.place(x=20, y=50)
        txt.place(x=300, y=80)
        progress(100, popupBar, bar, txt, 0, 0, "Maken totaaloverzicht... ")

        fileName = basename + "_totaal" + extension
        puoItem = 9
        puoList = get_list(content, puoItem)
        wb = xls.Workbook(fileName)
        datefmt = wb.add_format(get_dateformat())
        headerfmt = wb.add_format(get_headerformat())
        cellfmt = wb.add_format(get_cellformat())
        ws = wb.add_worksheet("Totaal")
        ws.set_column("A:A", 10, datefmt)

    header = ["Periode", "Foutcode", "Foutomschrijving"]
    for puo in puoList:
        header.append("# " + puo)
        header.append("#WGR " + puo)
        header.append("#IKV " + puo)
    header.append("# SWO's")
    header.append("Impact op proces")

    header.append("#Totaal")
    header.append("#Totaal wgr")
    header.append("#Totaal ikv")

    indexFields = "0,1,3,9"  # 'maand,fout' ##AANGEPAST OUD
    indexFields = "0,1,9"  # NIEUW
    currentCount = 0
    totalCount = len(content)
    write_worksheet(
        ws,
        content,
        cellfmt,
        indexFields,
        puoList,
        bar,
        txt,
        currentCount,
        totalCount,
        True,
    )

    wb.close()

    ibs.hide_window(popupBar)

    return


def process_swo(content, basename, extension):
    swoItem = 3
    puoItem = 9
    swoList = get_list(content, swoItem)
    puoList = get_list(content, puoItem)

    indexFields = "0,1,3,9"  # 'maand,fout,swo'

    dir = basename + "_per_SWO"
    posSlash = basename.rfind("/")
    cleanName = basename[posSlash:]
    if not os.path.isdir(dir):
        os.makedirs(dir)

    # progressbar
    ibs.show_window(popupBar)

    bar = ttk.Progressbar(
        popupBar, orient=tk.HORIZONTAL, length=950, mode="determinate"
    )
    txt = tk.Label(popupBar, text="Maken overzichten per SWO... ")
    bar.pack(expand=True)
    bar.place(x=20, y=50)
    txt.place(x=300, y=80)
    progress(100, popupBar, bar, txt, 0, 0, "Maken overzichten per SWO... ")
    currentCount = 0
    totalCount = 0

    for swo in swoList:
        contentIndex = make_index(content, indexFields)
        contentValue = fill_values(content, contentIndex, indexFields)
        totalCount += len(contentValue)

    for swo in swoList:
        if type(swo) is str:
            swo = swo.upper()
        fileName = dir + cleanName + "_" + str(swo) + extension
        swoContent = get_sub_content(content, swo, swoItem)
        wb = xls.Workbook(fileName)
        datefmt = wb.add_format(get_dateformat())
        headerfmt = wb.add_format(get_headerformat())
        cellfmt = wb.add_format(get_cellformat())
        ws = wb.add_worksheet(str(swo))
        ws.set_column("A:A", None, datefmt)

        header = ["Periode", "Foutcode", "Foutomschrijving"]
        for puo in puoList:
            header.append("# " + puo)
            header.append("#Totaal wgr " + puo)
            header.append("#Totaal ikv " + puo)
        header.append("Impact op proces")
        make_header(ws, ",".join(header), headerfmt)

        currentCount = write_worksheet(
            ws,
            swoContent,
            cellfmt,
            indexFields,
            puoList,
            bar,
            txt,
            currentCount,
            totalCount,
        )

        wb.close()

    ibs.hide_window(popupBar)

    return


def make_header(ws, headerNames, headerfmt):
    names = headerNames.split(",")
    col = 0
    for first, last, name in ibs.loop_with_first_last(names):
        if not first:
            width = 10
            if col == 2:
                width = 100
            if last:
                width = 200
            ws.set_column(col, col, width)
        ws.write(0, col, name, headerfmt)
        col += 1
    return names


def make_index(content, indexFields):
    contentIndex = []
    maxRows = len(content)
    for row in range(maxRows):
        indexKey = index_key(content, indexFields, row)
        if indexKey not in contentIndex:
            contentIndex.append(indexKey)
    contentIndex.sort()

    return contentIndex


def index_key(content, indexFields, row):
    keys = []
    fields = indexFields.split(",")
    # print(fields, 'helo')
    for field in fields:
        col = int(field)
        value = ibs.cell_value(content, row, col + 1)
        if type(value) is datetime:  # speciale behandeling voor de maand
            tmp = value.strftime("%y%m")
            value = tmp
        if type(value) is str:
            value = value.upper()
            value = value.strip("[")
            value = value.strip("]")
        keys.append(str(value))
    indexKey = "-".join(keys)

    return indexKey


def get_count(content, contentIndex, indexFields, countFields, indexSearch):
    count = 0

    values = count_values(content, contentIndex, indexFields, countFields, indexSearch)
    for row in values:
        if len(row) > 0:
            value = row[int(countFields)]
            if value is not None and type(value) == int:
                count += row[int(countFields)]

    return count


def get_impact(content, indexFields, impactCol, indexSearch):
    impact = ""
    maxRows = len(content)
    values = []

    for row in range(maxRows):
        indexKey = index_key(content, indexFields, row)
        if indexSearch is None or indexSearch == indexKey:
            value = ibs.cell_value(content, row, impactCol)
            if value is not None and value not in values:
                values.append(value)
    impact = "\n".join(values)

    return impact


def count_values(content, contentIndex, indexFields, countFields, indexSearch=None):
    contentValue = []
    fields = countFields.split(",")
    maxRows = len(content)
    maxCols = len(content[0]) - 1
    for item in range(len(contentIndex)):
        contentValue.append([])

    for row in range(maxRows):
        indexKey = index_key(content, indexFields, row)
        # print(contentValue)
        # print(indexSearch)
        if indexSearch is None or indexSearch == indexKey:
            indexNo = contentIndex.index(indexKey)
            # print(maxRows)
            if len(contentValue[indexNo]) == 0:
                # print(len(contentValue[indexNo]))
                for col in range(maxCols):
                    value = ibs.cell_value(
                        content, row, col + 1
                    )  # Nieuw: Dit lost het probleem bij regel 406 op.
                    # value = ibs.cell_value(content, row + 1, col + 1) #OUD
                    fields = [
                        int(x) for x in fields
                    ]  # 'fields' was een lijst van strings. 'col' is een int.
                    if (
                        col in fields
                    ):  # in de oude code was deze value nooit TRUE. Dus wanneer een foutcode maar 1 keer voor kwam in de maand werd deze nooit in de som meegenomen.
                        value = int(value)
                        # print(value, type(value))
                        contentValue[indexNo].append(int(value))
                    else:
                        contentValue[indexNo].append(value)
            else:
                for field in fields:
                    col = int(field)
                    value = ibs.cell_value(content, row, col + 1)  # Nieuw
                    # value = ibs.cell_value(content, row + 1, col + 1 #OUD
                    if value is not None:
                        contentValue[indexNo][col] = int(
                            contentValue[indexNo][col]
                        ) + int(value)
    return contentValue


def fill_values(content, contentIndex, indexFields):
    contentValue = []

    maxRows = len(content)
    maxCols = len(content[0])
    for item in range(len(contentIndex)):
        contentValue.append([])

    for row in range(maxRows):
        indexKey = index_key(content, indexFields, row)
        indexNo = contentIndex.index(indexKey)
        if len(contentValue[indexNo]) == 0:
            for col in range(1, maxCols + 1):
                value = ibs.cell_value(content, row, col)
                contentValue[indexNo].append(
                    value
                )  # hier worden van integers strings gemaakt!!!!!!
    return contentValue


def write_worksheet(
    ws,
    content,
    cellfmt,
    indexFields,
    puoList,
    bar,
    txt,
    currentCount,
    totalCount,
    total=False,
):
    swoItem = 3
    countItem = 5
    countWgItem = 6
    countIkvItem = 7
    impactItem = 8  # oud
    impactItem = 9  # nieuw
    # puoItem      = 9
    contentIndex = make_index(content, indexFields)
    contentValue = fill_values(
        content, contentIndex, indexFields
    )  # hier worden nog integers naar strings omgezet

    rowNo = 0
    countSWO = 0
    prevPerc = 0
    if total:
        processName = "Maken totaaloverzicht"
        totalCount = len(contentValue)
    else:
        processName = "Maken overzichten per SWO"

    for row in contentValue:
        rowNo += 1
        currentCount += 1
        prevPerc = progress(
            totalCount, popupBar, bar, txt, currentCount, prevPerc, processName
        )

        ws.write(rowNo, 0, row[0])  # periode
        ws.write(rowNo, 1, row[1])  # foutcode
        ws.write(rowNo, 2, row[2])  # foutomschrijving

        col = 3
        for puo in puoList:
            # dit intellen gaat fout!!!!! Er worden verkeerde waardes van verschillende keys bij elkaar opgeteld
            indexSearch = index_key(contentValue, indexFields, rowNo)
            countPuo = get_count(
                content, contentIndex, indexFields, str(countItem), indexSearch
            )
            countWgr = get_count(
                content, contentIndex, indexFields, str(countWgItem), indexSearch
            )
            countIkv = get_count(
                content, contentIndex, indexFields, str(countIkvItem), indexSearch
            )
            # print(indexSearch, countPuo, countWgr, countIkv)
            ws.write(rowNo, col, countPuo)
            ws.write(rowNo, col + 1, countWgr)
            ws.write(rowNo, col + 2, countIkv)
            col += 3

        if total:
            indexSearch = index_key(
                contentValue, indexFields + ",{}".format(swoItem), rowNo - 1
            )
            countSWO = get_count_index(content, indexSearch, indexFields + ",3")
            ws.write(rowNo, col, countSWO)  # aantal SWO's
            col += 1

        indexSearch = index_key(contentValue, indexFields, rowNo)
        impact = get_impact(content, indexFields, impactItem, indexSearch)
        ws.write(rowNo, col, impact, cellfmt)  # impact
        col += 1
        ws.write(rowNo, col, indexSearch)  # Later verwijderen

    return currentCount


def get_list(content, itemNo):
    list = []
    for row in content:
        item = row[itemNo]
        if not item in list:
            list.append(item)
    return list


def get_sub_content(content, sub, itemNo):
    subContent = []
    if type(sub) is str:
        sub = sub.upper()
    for item in content:
        check = item[itemNo]
        if type(check) is str:
            check = check.upper()
        if check == sub:
            subContent.append(item)

    return subContent


def get_count_index(content, indexSearch, indexFields):
    countIndex = 0
    maxRows = len(content)
    for row in range(maxRows):
        indexKey = index_key(content, indexFields, row)
        if indexSearch == indexKey:
            countIndex += 1
    return countIndex


def lees_bestand(filename):
    # progressbar
    ibs.show_window(popupBar)

    bar = ttk.Progressbar(
        popupBar, orient=tk.HORIZONTAL, length=950, mode="determinate"
    )
    txt = tk.Label(popupBar, text="Lezen excelbestanden... ")
    bar.pack(expand=True)
    bar.place(x=20, y=50)
    txt.place(x=300, y=80)
    progress(100, popupBar, bar, txt, 0, 0, "Lezen excelbestanden... ")

    wb = load_workbook(filename, read_only=True, data_only=True)
    names = wb.sheetnames
    name = names[0]
    wb.close()

    csvFile = ibs.excel_to_csv(filename, name)
    sheet = ibs.read_csv(csvFile)
    year = int(reportYear.get())
    prev = ""

    # lege waardes van de 1e 4 kolommen overnemen van een vorige waarde
    for row in sheet:
        for col in range(0, 3):
            if (row[col] == "" or row[col] is None) and prev != "":
                row[col] = prev[col]
        prev = row

    if sheet is not None:
        colMonth = -1
        colError = -1
        colDescription = -1
        colSWO = -1
        colName = -1
        colNoOfErrors = -1
        colNoOfWgr = -1
        colNoOfIkv = -1
        colImpact = -1
        maxContentCols = 9
        maxRows = ibs.get_max_rows(csvFile)
        maxCols = len(sheet[0])

        # aantal kolommen en rijen maken
        content = [[""] * (maxContentCols + 1) for i in range(maxRows - 1)]

        # kolommen bepalen
        for col in range(1, maxCols + 1):
            value = ibs.cell_value(sheet, 1, col, "upper")
            if value is not None:
                match value:
                    case "MAAND#":
                        colMonth = col
                    case "FOUTCODE":
                        colError = col
                    case "BESCHRIJVING FOUT":
                        colDescription = col
                    case "REL# SWO":
                        colSWO = col
                    case "NAAM SWO":
                        colName = col
                    case "Σ FOUT":
                        colNoOfErrors = col
                    case "Σ WGR":
                        colNoOfWgr = col
                    case "Σ IKV":
                        colNoOfIkv = col
                    case "BESCHRIJVING IMPACT":
                        colImpact = col

        repeatCol = [colMonth, colError, colDescription, colSWO, colName]

        # regels inlezen
        prevPerc = 0
        for row in range(2, maxRows + 1):  # eerste  rij overslaan, dat is de header
            prevPerc = progress(
                maxRows + 1,
                popupBar,
                bar,
                txt,
                row,
                prevPerc,
                "Lezen " + filename + ", tab " + name,
            )

            contentRow = row - 2
            content[contentRow][
                maxContentCols
            ] = name  # laatste col is bestemd voor de PUO

            for col in range(1, maxCols + 1):
                if col == colMonth:
                    content = fill_content(
                        sheet, content, repeatCol, row, col, 0, year, "month"
                    )
                if col == colError:
                    content = fill_content(
                        sheet, content, repeatCol, row, col, 1, year, "error"
                    )
                if col == colDescription:
                    content = fill_content(sheet, content, repeatCol, row, col, 2, year)
                if col == colSWO:
                    content = fill_content(
                        sheet, content, repeatCol, row, col, 3, year, "swo"
                    )
                if col == colName:
                    content = fill_content(sheet, content, repeatCol, row, col, 4, year)
                if col == colNoOfErrors:
                    content = fill_content(
                        sheet, content, repeatCol, row, col, 5, year, "int"
                    )
                if col == colNoOfWgr:
                    content = fill_content(
                        sheet, content, repeatCol, row, col, 6, year, "int"
                    )
                if col == colNoOfIkv:
                    content = fill_content(
                        sheet, content, repeatCol, row, col, 7, year, "int"
                    )
                if col == colImpact:
                    content = fill_content(sheet, content, repeatCol, row, col, 8, year)

    ibs.hide_window(popupBar)

    return content


def fill_content(ws, content, repeatCol, row, col, colContent, year, contentType=""):
    # month = 0
    contentRow = row - 2
    value = ibs.cell_value(ws, row, col)
    if (
        value is None or value == ""
    ) and col in repeatCol:  # niet gevuld, dan de vorige waarde gebruiken, indien opgegeven
        if contentRow > 0:
            value = content[contentRow - 1][colContent]  # vorige waarde
    match contentType:
        case "month":
            value = check_date(value, year)
            # if type(value) is str:
            #    month = ibs.monthnumber(value)
            # if month > 0:
            #    value = datetime(year, month, 1)
        case "error":
            try:
                strValue = str(value)
                chkDec = strValue.split(".")
                intValue = int(chkDec[0])
                if intValue < 10000:
                    value = "{:0>4d}".format(intValue)
                    if len(chkDec) > 1:
                        value = value + "." + chkDec[1]
                value = "[" + value + "]"
            except:
                if value is not None and len(value) > 1:
                    if value[:1].upper() == "P":
                        value = "[" + value + "]"
        case "swo":
            if value is not None and type(value) is str:
                value = value.upper()
        case "int":
            try:
                value = int(value)
            except:
                value = 0

    content[contentRow][colContent] = value

    return content


def check_date(value, year):
    chkDate = datetime
    if type(value) is not datetime:
        if value is not None:
            chkDate = ibs.string_to_date(value)
            if chkDate is None:
                month = ibs.monthnumber(value)
                if month == 0:
                    try:
                        month = int(value)
                    except:
                        pass
                if month != 0:
                    chkDate = datetime(year, month, 1)
    else:
        chkDate = value
    return chkDate


def clear_progress():
    ibs.clear_window(popupBar)


def progress(maxRows, window, bar, txt, row, prev, processName="Processed"):
    if maxRows > 0:
        percent = int((row / maxRows) * 100)
    if percent == 100 or percent > prev:
        percentText = f"{processName}  {percent} %"
        bar["value"] = percent
        txt["text"] = percentText
        if type(window) != "str":
            window.update_idletasks()
        prev = percent
    return prev


if __name__ == "__main__":
    root = ibs.create_root()
    main(root)
    root.mainloop()
